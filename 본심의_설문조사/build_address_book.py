# -*- coding: utf-8 -*-
"""
장비수요조사 주소록 엑셀 생성
- 출처1: '휴머노이드 AX 연구 프로젝트 협의체 모임' 시트 (PDF 기반)
- 출처2: 교수님께 추가 수령한 기업 담당자 리스트
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# (구분, 소속, 성명, 직위, 전화번호, 이메일, 전문분야, 참여여부, 비고)
rows = [
    # === 대학교수진 (협의체 시트) ===
    ('대학교수진', '서울대학교', '장병탁', '교수(원장)', '010-8647-7381', 'btzhang@bi.snu.ac.kr', '휴머노이드 학습/행동이론', 'O', ''),
    ('대학교수진', '서울대학교', '박재흥', '교수', '010-9077-2947', 'park73@snu.ac.kr', '휴머노이드 설계, 이족보행, 휴머노이드 행동, 자율주행', 'O', ''),
    ('대학교수진', '서울대학교', '최종현', '교수', '', 'jonghyunchoi@snu.ac.kr', '기계학습, 학습이론', 'O', ''),
    ('대학교수진', '서울대학교', '김건희', '교수', '', 'gunhee@snu.ac.kr', '컴퓨터 비전', 'X', ''),
    ('대학교수진', '서울대학교', '김현진', '교수', '', 'hjinkim@snu.ac.kr', '강화학습, UAV', 'O', ''),
    ('대학교수진', '서울대학교', '박용래', '교수', '', 'ylpark@snu.ac.kr', '촉각 센서, 소프트 로보틱스', 'O', ''),
    ('대학교수진', '서울대학교', '함종민', '교수(센터장)', '010-5012-0448', 'jmham@snu.ac.kr', '얼라이언스 연대', '', ''),
    ('대학교수진', '부산대학교', '이승준', '교수', '010-9042-4595', 'seungjoon.yi@pusan.ac.kr', '휴머노이드 설계, 휴머노이드 행동, 시스템 통합, 태스크 기반 실증', 'O', ''),
    ('대학교수진', 'POSTECH', '황인석', '교수', '', 'i.hwang@postech.ac.kr', '', 'O', ''),
    ('대학교수진', 'POSTECH', '조민수', '교수', '', 'mscho@postech.ac.kr', 'RFM', 'X', ''),
    ('대학교수진', 'KAIST', '박대형', '교수', '', 'daehyung@kaist.ac.kr', '휴머노이드 인지인식', 'O', ''),
    ('대학교수진', 'KAIST', '김범준', '교수', '', 'bumjoonkim@kaist.ac.kr', '로보틱스 전반, 3D Scene Understanding', 'O', ''),
    ('대학교수진', '고려대학교', '최성준', '교수', '010-5030-2735', 'sungjoon-choi@korea.ac.kr', '휴머노이드 모션학습', 'X', ''),
    ('대학교수진', '서강대학교', '남창주', '교수', '', 'cjnam@sogang.ac.kr', '로봇 지능, 로봇 모션 플래닝', 'X', ''),
    ('대학교수진', '연세대학교', '이영운', '교수', '010-6248-4928', 'youngwoon@yonsei.ac.kr', '강화학습, 로보틱스', 'X', ''),
    ('대학교수진', 'GIST', '이규빈', '교수', '010-3754-8909', 'kyoobinlee@gist.ac.kr', '시뮬레이션, 모션학습', '대리참석', ''),

    # === 기업 (협의체 시트) ===
    ('기업', '레인보우로보틱스', '오준호', '교수(CTO)', '010-5402-8049', '', '휴머노이드, 협동로봇', '', ''),
    ('기업', '레인보우로보틱스', '김인혁', '부사장', '010-8999-1342', 'inhyeok.kim@rainbow-robotics.com', '', '', ''),
    ('기업', '레인보우로보틱스', '이정호', '대표', '', '', '', '', ''),
    ('기업', '레인보우로보틱스', '허정우', '기술이사', '010-9908-8919', 'jwheo@rainbow-robotics.com', '', '', '사무실 042-719-8070'),
    ('기업', '에이로봇', '한재권', '교수(CTO)', '010-4186-2802', 'jkhan@hanyang.ac.kr', '휴머노이드', '', ''),
    ('기업', '에이로봇', '엄윤설', '대표', '010-6292-2802', 'shealeum@arobot4all.com', '', 'O', ''),
    ('기업', '홀리데이로보틱스', '송기영', '대표', '010-2589-9744', 'kiyoung.song@holiday-robotics.com', '휴머노이드(상체), 시뮬레이션', 'O', ''),
    ('기업', '위로보틱스', '김용재', '교수(대표)', '010-8820-6360', 'yongjae@koreatech.ac.kr', '휴머노이드, 웨어러블 로봇', '', '한국기술교육대학교 교수 겸임'),
    ('기업', '딥엑스', '김정욱', '부사장', '010-9003-0731', 'jkim@deepx.ai', 'NPU', '', ''),
    ('기업', '딥엑스', '정한별', '이사', '010-9930-1791', '', '', '', ''),
    ('기업', '투모로', '장준현', '부대표', '010-6369-2890', 'jhjang@tommoro.ai', 'RFM 제작', 'O', ''),

    # === 기업 (교수님 추가분) ===
    # 원익로보틱스
    ('기업', '원익로보틱스', '김민철', '전무', '010-5156-3522', 'mckim@wonik.com', '', '', ''),
    ('기업', '원익로보틱스', '장성진', '상무', '010-8759-7807', 'sj.jang@wonik.com', '', '', ''),
    ('기업', '원익로보틱스', '송창우', '팀장', '010-5037-0143', 'chw_song@wonik.com', '', '', ''),
    ('기업', '원익로보틱스', '정인호', '부장', '010-3182-3724', 'inho.jung@wonik.com', '', '', ''),
    ('기업', '원익로보틱스', '변상철', '부장', '010-4169-8284', 'scbyun@wonik.com', '', '', ''),
    ('기업', '원익로보틱스', '김대현', '전무', '010-7189-8360', 'tonykim@wonik.com', '로봇 손 제작', 'O', ''),
    ('기업', '원익로보틱스', '김현수', '팀장', '010-2385-5489', 'hskim566@wonik.com', '', '', ''),
    ('기업', '원익로보틱스', '서광윤', '팀장', '010-9441-3333', 'kyseo@wonik.com', '', '', ''),
    # 플라잎
    ('기업', '플라잎', '정태영', '대표', '010-2795-0881', 'jty@plaif.com', '', '', ''),
    ('기업', '플라잎', '이주성', '그룹장', '010-6876-5211', 'jslee@plaif.com', '', '', ''),
    ('기업', '플라잎', '박은섭', '', '010-9644-3606', 'espark@plaif.com', '', '', ''),
    # 기타 단일 인원 기업
    ('기업', '건솔루션', '정일화', '센터장', '010-7577-7897', 'jih@gunsol.com', '', '', ''),
    ('기업', '두산로보틱스', '김민표', '대표', '010-4302-4057', 'minpyo.kim@doosan.com', '', '', ''),
    ('기업', '티로보틱스', '심영보', '본부장', '010-5447-6882', 'ybshim@t-robotics.co.kr', '', '', ''),
    ('기업', '코라로보틱스', '송재복', '대표', '010-3726-3363', 'jbsong@korarobotics.com', '', '', ''),
    ('기업', '유일로보틱스', '한정헌', '전무', '010-8901-5619', 'jhan@yuilrobotics.com', '', '', ''),
    ('기업', '티라로보틱스', '김동경', '대표', '010-8983-8290', 'pdongkim@thirarobotics.com', '', '', ''),
    ('기업', '에이딘로보틱스', '이윤행', '대표', '010-8955-4411', 'ymove@aidinrobotics.com', '', '', ''),
    ('기업', '휴민로보틱스', '서종휘', '대표', '010-2279-4191', 'jongwi.seo@ihumin.co.kr', '', '', ''),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = '장비수요조사_주소록'

    headers = ['연번', '구분', '소속', '성명', '직위', '전화번호', '이메일',
               '발송일', '회신여부', '회신일', '전화 재확인일']

    # 헤더
    header_font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='305496')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='A0A0A0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # 데이터
    body_font = Font(name='맑은 고딕', size=10)
    alt_fill = PatternFill('solid', fgColor='F2F6FC')

    for idx, r in enumerate(rows, start=1):
        row_num = idx + 1
        # r = (구분, 소속, 성명, 직위, 전화번호, 이메일, 전문분야, 참여여부, 비고)
        # 사용 컬럼: 구분, 소속, 성명, 직위, 전화번호, 이메일만
        base = list(r[:6])
        vals = [idx] + base + ['', '', '', '']
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.font = body_font
            c.border = border
            # 중앙 정렬: 연번, 구분, 발송일, 회신여부, 회신일, 전화 재확인일
            if col in (1, 2, 8, 9, 10, 11):
                c.alignment = center
            else:
                c.alignment = left

    # 열 너비
    widths = [5, 10, 15, 8, 12, 14, 28, 12, 10, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 행 높이 기본
    ws.row_dimensions[1].height = 30

    # 틀 고정
    ws.freeze_panes = 'A2'

    # 자동 필터
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(rows) + 1}'

    out = os.path.join(os.path.dirname(__file__), '장비수요조사_주소록_v1.xlsx')
    wb.save(out)
    print(f'생성 완료: {out}')
    print(f'총 {len(rows)}명 등록')


if __name__ == '__main__':
    build()
